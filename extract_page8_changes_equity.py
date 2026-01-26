import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class BudimexPage8Extractor:
    """Extract Consolidated Statement of Changes in Equity (Page 8) with exact values"""
    
    def __init__(self, pdf_path: str, output_path: str = "Budimex_Page8_Changes_in_Equity.xlsx"):
        self.pdf_path = pdf_path
        self.output_path = output_path
        self.extracted_data = []
        self.headers = []
    
    def extract_page_8(self):
        """Extract page 8 with all exact values and structure"""
        print(f"🔄 Opening PDF: {self.pdf_path}")
        
        with pdfplumber.open(self.pdf_path) as pdf:
            print(f"📄 Total pages: {len(pdf.pages)}")
            
            # Extract page 8 (0-based indexing = page 7)
            page_8 = pdf.pages[7]
            tables = page_8.extract_tables()
            
            if not tables:
                print("❌ No tables found on page 8")
                return False
            
            print(f"📊 Found {len(tables)} table(s) on page 8")
            
            # Process each table on page 8
            for table_idx, table in enumerate(tables):
                print(f"\n📋 Processing Table {table_idx + 1}")
                print(f"   Rows: {len(table)}, Columns: {len(table[0]) if table else 0}")
                
                # Extract headers (first row)
                if table:
                    self.headers = table[0]
                    print(f"   Headers: {len(self.headers)} columns")
                    
                    # Extract data rows (skip header)
                    data_rows = table[1:]
                    
                    for row_idx, row in enumerate(data_rows):
                        # Clean and pad row
                        cleaned_row = []
                        for cell in row:
                            if cell is None:
                                cleaned_row.append("")
                            else:
                                cleaned_row.append(str(cell).strip())
                        
                        # Pad with empty strings if needed
                        while len(cleaned_row) < len(self.headers):
                            cleaned_row.append("")
                        
                        # Keep only the columns we have headers for
                        cleaned_row = cleaned_row[:len(self.headers)]
                        
                        self.extracted_data.append(cleaned_row)
                    
                    print(f"   ✅ Extracted {len(self.extracted_data)} data rows")
        
        return True if self.extracted_data else False
    
    def create_excel_file(self):
        """Create Excel file with exact table structure"""
        if not self.extracted_data or not self.headers:
            print("❌ No data to export")
            return False
        
        print(f"\n📝 Creating Excel file: {self.output_path}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Changes_in_Equity"
        
        # Define styles
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        
        subtotal_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        subtotal_font = Font(bold=True, color="000000", size=10)
        
        total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        total_font = Font(bold=True, color="000000", size=11)
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # Write header row (row 1)
        for col_idx, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Write data rows
        for row_idx, row_data in enumerate(self.extracted_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value if value else ""
                cell.border = border
                
                # Determine alignment based on content
                if col_idx == 1:  # First column (row labels)
                    cell.alignment = left_align
                    # Bold for subtotals and totals
                    row_label = value.lower() if value else ""
                    if any(keyword in row_label for keyword in ['balance', 'comprehensive', 'payment', 'contribution', 'sale']):
                        cell.font = subtotal_font
                else:  # Numeric columns
                    cell.alignment = right_align
                    # Try to format as number
                    try:
                        # Remove parentheses for negative numbers
                        if isinstance(value, str):
                            if value.startswith('(') and value.endswith(')'):
                                numeric_val = float(value[1:-1].replace(' ', ''))
                                cell.value = -numeric_val
                            else:
                                # Try to parse number
                                numeric_val = float(value.replace(',', '').replace(' ', ''))
                                cell.value = numeric_val
                            cell.number_format = '#,##0'
                    except (ValueError, AttributeError):
                        pass
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(self.headers, 1):
            max_length = len(str(header))
            
            # Check data rows for max length
            for row_data in self.extracted_data:
                if col_idx <= len(row_data):
                    cell_length = len(str(row_data[col_idx - 1]))
                    max_length = max(max_length, cell_length)
            
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Set row heights for better visibility
        ws.row_dimensions[1].height = 25  # Header row
        
        for row_idx, row_data in enumerate(self.extracted_data, 2):
            row_label = row_data[0].lower() if row_data else ""
            if any(keyword in row_label for keyword in ['balance as at', 'comprehensive income']):
                ws.row_dimensions[row_idx].height = 18
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(self.output_path)
        print(f"✅ Excel file saved: {self.output_path}")
        print(f"   - Rows: {len(self.extracted_data)} (+ 1 header)")
        print(f"   - Columns: {len(self.headers)}")
        
        return True


def main():
    pdf_file = "2024_Budimex.pdf"
    output_file = "Budimex_Page8_Changes_in_Equity.xlsx"
    
    if not os.path.exists(pdf_file):
        print(f"❌ PDF file not found: {pdf_file}")
        print(f"   Expected: {os.path.abspath(pdf_file)}")
        return
    
    # Initialize extractor
    extractor = BudimexPage8Extractor(pdf_file, output_file)
    
    # Extract page 8
    print("\n🚀 Starting Page 8 Extraction...\n")
    if extractor.extract_page_8():
        # Create Excel file
        if extractor.create_excel_file():
            print(f"\n✨ SUCCESS! File created: {output_file}")
    else:
        print("\n❌ Extraction failed - no data found")


if __name__ == "__main__":
    main()
