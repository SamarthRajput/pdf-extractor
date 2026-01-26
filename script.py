import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class FinancialStatementExtractorSingleSheet:
    
    def __init__(self, pdf_path: str, output_path: str = "financial_statements.xlsx"):
        self.pdf_path = pdf_path
        self.output_path = output_path
        self.extracted_tables = []
        self.table_metadata = []
        
        self.standard_columns = [
            'primary_key', 'date_last_updated', 'published_date', 'reported_date',
            'doc_page_num', 'file_page_num', 'table_id', 'country', 'geo_1_id',
            'geo_1_name', 'geo_1_type', 'geo_2_id', 'geo_2_name', 'geo_2_type',
            'dim_4_id', 'dim_4_name', 'dim_3_id', 'dim_3_name', 'dim_2_id',
            'dim_2_name', 'dim_1_id', 'dim_1_name', 'metric_id', 'metric_name',
            'source_metric_id', 'source_metric_name', 'indentation', 'process_flag',
            'base_factor', 'display_power_factor', 'data_frequency', 'aggregation_method',
            'unit', 'unit_type', 'note_id', 'note_reference', 'cumulative_periods',
            'comments', 'check_sum', 'concat', 'formula', '2022', '2023', '2024',
            '2022_check', '2023_check', '2024_check'
        ]
    
    def extract_page_8_only(self):
        """Extract page 8 using text extraction with coordinates"""
        print(f"🔄 Opening PDF: {self.pdf_path}")
        print(f"📄 Extracting Page 8 (Changes in Equity)")
        
        with pdfplumber.open(self.pdf_path) as pdf:
            if len(pdf.pages) < 8:
                print(f"❌ PDF has only {len(pdf.pages)} pages")
                return self.extracted_tables
            
            page = pdf.pages[7]  # Page 8 (0-indexed)
            
            # Extract text with positions
            words = page.extract_words()
            
            # Group words by vertical position (rows)
            rows_dict = {}
            for word in words:
                y_pos = round(word['top'])
                if y_pos not in rows_dict:
                    rows_dict[y_pos] = []
                rows_dict[y_pos].append(word)
            
            # Sort rows by vertical position
            sorted_rows = sorted(rows_dict.items(), key=lambda x: x[0])
            
            # Build table rows
            table_data = []
            for y_pos, words_in_row in sorted_rows:
                # Sort words by horizontal position
                sorted_words = sorted(words_in_row, key=lambda w: w['x0'])
                
                # Join words into cells based on position
                row_text = ' '.join([w['text'] for w in sorted_words])
                
                # Skip empty rows
                if row_text.strip():
                    table_data.append(sorted_words)
            
            # Convert to DataFrame (simplified approach)
            if table_data:
                # Take first 15 rows (approximate table size)
                table_subset = table_data[:15]
                
                # Extract as simple rows
                simple_rows = []
                for row_words in table_subset:
                    row_text = [w['text'] for w in row_words]
                    simple_rows.append(row_text)
                
                # Create DataFrame (first row as headers)
                if simple_rows:
                    headers = simple_rows[0]
                    data_rows = simple_rows[1:]
                    
                    # Pad rows to match header length
                    max_len = len(headers)
                    padded_rows = []
                    for row in data_rows:
                        while len(row) < max_len:
                            row.append("")
                        padded_rows.append(row[:max_len])
                    
                    df = pd.DataFrame(padded_rows, columns=headers)
                    
                    self.extracted_tables.append(df)
                    self.table_metadata.append({
                        'page': 8,
                        'table_index': 0,
                        'rows': len(df),
                        'columns': len(df.columns),
                        'title': "Changes in Equity"
                    })
                    
                    print(f"✅ Extracted {len(df)} rows × {len(df.columns)} columns")
        
        return self.extracted_tables
    
    def extract_all_tables(self):
        """Extract all tables from entire PDF"""
        print(f"🔄 Opening PDF: {self.pdf_path}")
        
        with pdfplumber.open(self.pdf_path) as pdf:
            print(f"📄 Total pages: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()
                
                if tables:
                    print(f"📊 Found {len(tables)} table(s) on page {page_num}")
                    
                    for table_idx, table in enumerate(tables):
                        df = self._convert_table_to_dataframe(table)
                        
                        if df is not None and not df.empty:
                            self.extracted_tables.append(df)
                            self.table_metadata.append({
                                'page': page_num,
                                'table_index': table_idx,
                                'rows': len(df),
                                'columns': len(df.columns)
                            })
                            print(f"✅ Table {len(self.extracted_tables)}: Page {page_num}, {len(df)} rows × {len(df.columns)} columns")
        
        return self.extracted_tables
    
    def _convert_table_to_dataframe(self, table):
        """Convert raw table to DataFrame"""
        if not table or len(table) < 2:
            return None
        
        try:
            headers = table[0]
            data_rows = table[1:]
            data_rows = [row for row in data_rows if any(cell for cell in row)]
            
            if not data_rows:
                return None
            
            max_cols = len(headers)
            padded_rows = []
            for row in data_rows:
                while len(row) < max_cols:
                    row.append("")
                padded_rows.append(row[:max_cols])
            
            df = pd.DataFrame(padded_rows, columns=headers)
            df = self._clean_dataframe(df)
            
            return df
        except Exception as e:
            print(f"⚠️ Error converting table: {e}")
            return None
    
    def _clean_dataframe(self, df):
        """Clean DataFrame"""
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
        df = df.replace('', None)
        return df.reset_index(drop=True)
    
    def convert_numeric_columns(self):
        """Convert numeric columns"""
        for df in self.extracted_tables:
            for col in df.columns:
                df[col] = df[col].apply(lambda x: self._convert_to_numeric(x))
    
    # def _convert_to_numeric(self, value):
    #     """Convert value to numeric: remove spaces/commas, handle parentheses"""
    #     if value is None:
    #         return None
        
    #     if isinstance(value, pd.Series):
    #         return value
        
    #     value_str = str(value).strip()
        
    #     if value_str == "" or value_str.lower() == "nan":
    #         return None
        
    #     if value_str == "-" or value_str == "–":
    #         return None
        
    #     # Handle parentheses: (123) -> -123
    #     if value_str.startswith('(') and value_str.endswith(')'):
    #         try:
    #             inner = value_str[1:-1].replace(',', '').replace(' ', '')
    #             num = float(inner)
    #             return -num
    #         except:
    #             return value_str
        
    #     # Remove commas and spaces
    #     try:
    #         clean_value = value_str.replace(',', '').replace(' ', '')
    #         num = float(clean_value)
    #         return num
    #     except:
    #         return value_str
        
    def _convert_to_numeric(self, value):
        """
        Convert value to numeric with specific rules:
        - Remove all spaces: "1 20 300" -> 120300
        - Remove all commas: "1,23,400" -> 123400
        - Convert parentheses to negative: "(123)" -> -123
        """
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return value
        
        if isinstance(value, pd.Series):
            return value
        
        value_str = str(value).strip()
        
        # Empty or nan
        if value_str == "" or value_str.lower() == "nan":
            return None
        
        # Just a dash/hyphen means null
        if value_str == "-" or value_str == "–":
            return None
        
        # Handle parentheses for negative: (123) -> -123
        # Remove spaces and commas first
        if value_str.startswith('(') and value_str.endswith(')'):
            try:
                inner = value_str[1:-1]
                # Remove ALL spaces and commas
                inner = inner.replace(',', '').replace(' ', '').replace('\xa0', '')
                num = float(inner)
                return -num  # Return as numeric, not string
            except:
                return value_str
        
        # Try to convert regular number
        try:
            # Remove ALL commas, spaces, and non-breaking spaces
            clean_value = value_str.replace(',', '').replace(' ', '').replace('\xa0', '')
            # Try to parse as float
            num = float(clean_value)
            print("fixed values are " + num)
            return num  # Return as numeric, not string
        except:
            # Not a number, return as-is (text)
            return value_str

    def transform_to_standard_schema(self):
        """Transform to standard schema"""
        print(f"\n📝 Transforming {len(self.extracted_tables)} table(s) to standardized schema...")
        
        transformed_data = []
        primary_key_counter = 1
        
        for table_idx, (df, metadata) in enumerate(zip(self.extracted_tables, self.table_metadata)):
            print(f"\n📊 Processing Table {table_idx + 1} (Page {metadata['page']})")
            
            for row_idx, row in enumerate(df.itertuples(index=False)):
                row_data = {}
                row_data['primary_key'] = primary_key_counter
                row_data['date_last_updated'] = datetime.now().strftime('%Y-%m-%d')
                row_data['published_date'] = datetime.now().strftime('%Y-%m-%d')
                row_data['reported_date'] = '2024-12-31'
                row_data['doc_page_num'] = metadata['page']
                row_data['file_page_num'] = metadata['page']
                row_data['table_id'] = f"TABLE_{table_idx + 1:03d}"
                row_data['country'] = 'PL'
                
                for field in ['geo_1_id', 'geo_1_name', 'geo_1_type', 'geo_2_id', 'geo_2_name', 'geo_2_type']:
                    row_data[field] = None
                
                extracted_cols = list(row)
                
                for i in range(1, 5):
                    if len(extracted_cols) >= i:
                        row_data[f'dim_{i}_id'] = f"DIM{i}_{table_idx}_{row_idx}"
                        row_data[f'dim_{i}_name'] = str(extracted_cols[i-1])
                    else:
                        row_data[f'dim_{i}_id'] = None
                        row_data[f'dim_{i}_name'] = None
                
                row_data['metric_id'] = f"METRIC_{table_idx}_{row_idx}"
                row_data['metric_name'] = f"Value_{row_idx}"
                row_data['source_metric_id'] = f"SOURCE_{table_idx}_{row_idx}"
                row_data['source_metric_name'] = "PDF_Extracted"
                row_data['indentation'] = 0
                row_data['process_flag'] = 1
                row_data['base_factor'] = 1
                row_data['display_power_factor'] = 0
                row_data['data_frequency'] = 'Annual'
                row_data['aggregation_method'] = 'Sum'
                row_data['unit'] = 'PLN'
                row_data['unit_type'] = 'Thousands'
                row_data['note_id'] = None
                row_data['note_reference'] = None
                row_data['cumulative_periods'] = 1
                row_data['comments'] = f"Extracted from page {metadata['page']}"
                
                for year in ['2022', '2023', '2024']:
                    row_data[year] = None
                    row_data[f'{year}_check'] = None
                
                if len(extracted_cols) >= 4:
                    row_data['2024'] = extracted_cols[-1]
                if len(extracted_cols) >= 5:
                    row_data['2023'] = extracted_cols[-2]
                if len(extracted_cols) >= 6:
                    row_data['2022'] = extracted_cols[-3]
                
                row_data['check_sum'] = None
                row_data['concat'] = None
                row_data['formula'] = None
                
                transformed_data.append(row_data)
                primary_key_counter += 1
        
        result_df = pd.DataFrame(transformed_data, columns=self.standard_columns)
        print(f"✅ Transformed {len(result_df)} rows")
        return result_df
    
    # def create_single_sheet_excel(self, transformed_df):
    #     """Create Excel file"""
    #     print(f"\n📝 Creating: {self.output_path}")
        
    #     wb = Workbook()
    #     ws = wb.active
    #     ws.title = "Consolidated_Data"
        
    #     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    #     header_font = Font(bold=True, color="FFFFFF", size=10)
    #     border = Border(
    #         left=Side(style='thin'),
    #         right=Side(style='thin'),
    #         top=Side(style='thin'),
    #         bottom=Side(style='thin')
    #     )
        
    #     for col_idx, header in enumerate(transformed_df.columns, 1):
    #         cell = ws.cell(row=1, column=col_idx)
    #         cell.value = header
    #         cell.fill = header_fill
    #         cell.font = header_font
    #         cell.border = border
    #         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    #     for row_idx, row in enumerate(transformed_df.values, 2):
    #         for col_idx, value in enumerate(row, 1):
    #             cell = ws.cell(row=row_idx, column=col_idx)
    #             cell.value = value
    #             cell.border = border
                
    #             col_name = transformed_df.columns[col_idx - 1]
    #             if col_name in ['2022', '2023', '2024', '2022_check', '2023_check', '2024_check']:
    #                 cell.alignment = Alignment(horizontal='right', vertical='center')
    #                 if isinstance(value, (int, float)):
    #                     cell.number_format = '#,##0'
    #             else:
    #                 cell.alignment = Alignment(horizontal='left', vertical='center')
        
    #     for col_idx in range(1, len(transformed_df.columns) + 1):
    #         ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
    #     ws.freeze_panes = 'A2'
    #     wb.save(self.output_path)
    #     print(f"✅ Saved: {self.output_path}")
    #     return self.output_path

    def create_single_sheet_excel(self, transformed_df):
        """Create Excel file with all data in a SINGLE SHEET"""
        print(f"\n📝 Creating Excel file with single sheet: {self.output_path}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated_Data"
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Write headers
        for col_idx, header in enumerate(transformed_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Write data
        for row_idx, row in enumerate(transformed_df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Get column name
                col_name = transformed_df.columns[col_idx - 1]
                
                # Clean and convert numeric values before writing
                if col_name in ['2022', '2023', '2024', '2022_check', '2023_check', '2024_check', 
                            'base_factor', 'display_power_factor']:
                    # Apply conversion again to ensure it's numeric
                    converted_value = self._convert_to_numeric(value)
                    cell.value = converted_value
                    cell.alignment = right_align
                    
                    # Apply number format ONLY if it's actually a number
                    if isinstance(converted_value, (int, float)) and converted_value is not None:
                        cell.number_format = '#,##0'
                else:
                    cell.value = value
                    cell.alignment = left_align
                
                cell.border = border
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(transformed_df.columns, 1):
            col_name = transformed_df.columns[col_idx - 1]
            if col_name in ['2022', '2023', '2024', '2022_check', '2023_check', '2024_check']:
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            elif col_name in ['primary_key', 'table_id', 'country']:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        wb.save(self.output_path)
        print(f"✅ Excel file saved: {self.output_path}")
        print(f"📊 Total rows in sheet: {len(transformed_df)}")
        print(f"📊 Total columns: {len(transformed_df.columns)}")
        
        return self.output_path



def main_option_1_page_8_only():
    pdf_file = "2024_Budimex.pdf"
    output_file = "Budimex_Page8_StandardSchema.xlsx"
    
    if not os.path.exists(pdf_file):
        print(f"❌ PDF not found: {pdf_file}")
        return
    
    extractor = FinancialStatementExtractorSingleSheet(pdf_file, output_file)
    extractor.extract_page_8_only()
    extractor.convert_numeric_columns()
    transformed_df = extractor.transform_to_standard_schema()
    extractor.create_single_sheet_excel(transformed_df)
    print(f"\n✨ Complete! Output: {output_file}")


def main_option_2_whole_pdf():
    pdf_file = "2024_Budimex.pdf"
    output_file = "Budimex_WholePDF_StandardSchema.xlsx"
    
    if not os.path.exists(pdf_file):
        print(f"❌ PDF not found: {pdf_file}")
        return
    
    extractor = FinancialStatementExtractorSingleSheet(pdf_file, output_file)
    extractor.extract_all_tables()
    extractor.convert_numeric_columns()
    transformed_df = extractor.transform_to_standard_schema()
    extractor.create_single_sheet_excel(transformed_df)
    print(f"\n✨ Complete!")


if __name__ == "__main__":
    # main_option_1_page_8_only()
    main_option_2_whole_pdf()
