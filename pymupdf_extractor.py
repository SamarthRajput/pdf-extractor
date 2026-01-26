import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class BudimexPage8Extractor:
    
    def __init__(self, pdf_path, output_path="Budimex_Page8.xlsx"):
        self.pdf_path = pdf_path
        self.output_path = output_path
        self.table_data = []

    def extract_page_8(self):
        if not os.path.exists(self.pdf_path):
            print(f"❌ File not found: {self.pdf_path}")
            return False

        print(f"🔄 Opening: {self.pdf_path}")
        
        with pdfplumber.open(self.pdf_path) as pdf:
            if len(pdf.pages) < 8:
                print(f"❌ Only {len(pdf.pages)} pages")
                return False

            page = pdf.pages[7]  # Page 8
            
            # Use explicit table settings to detect columns properly
            table_settings = {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "join_tolerance": 3,
                "edge_min_length": 3,
                "min_words_vertical": 1,
                "intersection_tolerance": 3
            }
            
            # Extract table with custom settings
            table = page.extract_table(table_settings)
            
            if not table:
                print("❌ No table found")
                return False
            
            print(f"📊 Extracted table: {len(table)} rows")
            
            # Store all rows
            self.table_data = []
            for row in table:
                cleaned_row = []
                for cell in row:
                    if cell is None or cell == "":
                        cleaned_row.append("")
                    else:
                        cleaned = str(cell).replace('\n', ' ').strip()
                        cleaned_row.append(cleaned)
                self.table_data.append(cleaned_row)
            
            # Remove completely empty rows
            self.table_data = [row for row in self.table_data if any(cell for cell in row)]
            
            print(f"✅ Parsed {len(self.table_data)} rows")
            
            if self.table_data:
                print(f"   Columns: {len(self.table_data[0])}")
            
            return True

    def _parse_number(self, value):
        if not value or value == "-":
            return None
        text = str(value).strip()
        try:
            text = text.replace('\n', '').replace('\r', '')
            if text.startswith("(") and text.endswith(")"):
                return -float(text[1:-1].replace(",", "").replace(" ", ""))
            return float(text.replace(",", "").replace(" ", ""))
        except:
            return text

    def create_excel(self):
        if not self.table_data:
            print("❌ No data")
            return False

        print(f"\n📝 Creating: {self.output_path}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Page_8"

        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        header_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        total_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row_idx, row in enumerate(self.table_data, 1):
            is_header = row_idx == 1
            is_total = any(keyword in str(row[0]).lower() for keyword in ['balance as at', 'balance at']) if row else False
            
            for col_idx, val in enumerate(row, 1):
                c = ws.cell(row_idx, col_idx)
                
                if is_header:
                    c.value = val
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif col_idx == 1:
                    c.value = val
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if is_total:
                        c.font = total_font
                        c.fill = total_fill
                else:
                    parsed = self._parse_number(val)
                    c.value = parsed
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(parsed, (int, float)):
                        c.number_format = "#,##0"
                    if is_total:
                        c.font = total_font
                        c.fill = total_fill
                
                c.border = border

        for col_idx in range(1, len(self.table_data[0]) + 1 if self.table_data else 1):
            max_len = 10
            for row in self.table_data:
                if col_idx <= len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        ws.freeze_panes = "A2"
        wb.save(self.output_path)
        print(f"✅ Saved: {self.output_path}")
        print(f"   {len(self.table_data)} rows × {len(self.table_data[0]) if self.table_data else 0} columns")
        return True


def main():
    extractor = BudimexPage8Extractor("2024_Budimex.pdf")
    if extractor.extract_page_8():
        extractor.create_excel()


if __name__ == "__main__":
    main()
