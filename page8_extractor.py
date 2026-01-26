import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class BudimexPage8Extractor:
    """Extract Consolidated Statement of Changes in Equity (Page 8)"""
    
    def __init__(self, pdf_path: str, output_path: str = "Budimex_Page8_Changes_in_Equity.xlsx"):
        self.pdf_path = pdf_path
        self.output_path = output_path
        self.headers = []
        self.data_rows = []
    
    def extract_page_8(self):
        """Extract page 8 table with exact values"""
        print(f"🔄 Opening PDF: {self.pdf_path}")
        
        with pdfplumber.open(self.pdf_path) as pdf:
            print(f"📄 Total pages: {len(pdf.pages)}")
            
            if len(pdf.pages) < 8:
                print(f"❌ PDF has only {len(pdf.pages)} pages")
                return False
            
            # Extract page 8 (index 7)
            page = pdf.pages[7]
            tables = page.extract_tables()
            
            if not tables:
                print("❌ No tables found on page 8")
                return False
            
            print(f"📊 Found {len(tables)} table(s) on page 8")
            
            # Process first table (main statement)
            table = tables[0]
            
            print(f"   Table structure: {len(table)} rows × {len(table[0]) if table else 0} cols")
            
            # First row is headers
            if table:
                self.headers = [str(h).strip() if h else "" for h in table[0]]
                print(f"   Headers: {len(self.headers)} columns")
                
                # All other rows are data
                for row_idx, row in enumerate(table[1:], 1):
                    # Convert row to strings and handle None values
                    cleaned_row = []
                    for cell in row:
                        if cell is None:
                            cleaned_row.append("")
                        else:
                            cleaned_row.append(str(cell).strip())
                    
                    # Pad with empty strings to match header count
                    while len(cleaned_row) < len(self.headers):
                        cleaned_row.append("")
                    
                    # Trim excess columns
                    cleaned_row = cleaned_row[:len(self.headers)]
                    
                    self.data_rows.append(cleaned_row)
                
                print(f"   ✅ Extracted {len(self.data_rows)} data rows")
                return True
        
        return False
    
    def _is_total_or_subtotal(self, cell_value):
        """Check if row is a total or subtotal"""
        if not cell_value:
            return False
        text = str(cell_value).lower()
        keywords = ['balance', 'comprehensive', 'payment', 'contribution', 'sale']
        return any(keyword in text for keyword in keywords)
    
    def _try_parse_number(self, value):
        """Try to parse a value as a number"""
        if not value or not isinstance(value, str):
            return value
        
        text = value.strip()
        
        # Handle empty
        if text == "" or text == "-":
            return None
        
        try:
            # Handle negative values in parentheses: (123) -> -123
            if text.startswith('(') and text.endswith(')'):
                num_str = text[1:-1].replace(',', '').replace(' ', '')
                return float(num_str) * -1
            
            # Handle regular numbers with thousand separators
            num_str = text.replace(',', '').replace(' ', '')
            return float(num_str)
        except (ValueError, AttributeError):
            return text
    
    def create_excel_file(self):
        """Create Excel file with proper formatting"""
        if not self.headers or not self.data_rows:
            print("❌ No data to export")
            return False
        
        print(f"\n📝 Creating Excel file: {self.output_path}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Changes_in_Equity"
        
        # Define styles
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        
        total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        total_font = Font(bold=True, color="000000", size=10)
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Write header row
        for col_idx, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data rows
        for row_idx, row_data in enumerate(self.data_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # First column: text (row labels)
                if col_idx == 1:
                    cell.value = value
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = border
                    
                    # Bold if total/subtotal
                    if self._is_total_or_subtotal(value):
                        cell.font = total_font
                        cell.fill = total_fill
                else:
                    # Other columns: try to parse as numbers
                    parsed_value = self._try_parse_number(value)
                    cell.value = parsed_value
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border
                    
                    # Format numbers
                    if isinstance(parsed_value, (int, float)):
                        cell.number_format = '#,##0'
                    
                    # Bold if total/subtotal
                    if self._is_total_or_subtotal(row_data[0]):
                        cell.font = total_font
                        cell.fill = total_fill
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(self.headers, 1):
            max_length = len(str(header))
            
            for row_data in self.data_rows:
                if col_idx <= len(row_data):
                    cell_len = len(str(row_data[col_idx - 1]))
                    max_length = max(max_length, cell_len)
            
            adjusted_width = min(max_length + 3, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Set header row height
        ws.row_dimensions[1].height = 30
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(self.output_path)
        
        print(f"✅ Excel file saved: {self.output_path}")
        print(f"   📊 Data: {len(self.data_rows)} rows × {len(self.headers)} columns")
        
        return True


def main():
    pdf_file = "2024_Budimex.pdf"
    output_file = "Budimex_Page8_Changes_in_Equity.xlsx"
    
    if not os.path.exists(pdf_file):
        print(f"❌ Error: {pdf_file} not found")
        return
    
    extractor = BudimexPage8Extractor(pdf_file, output_file)
    
    print("\n🚀 EXTRACTING PAGE 8: Consolidated Statement of Changes in Equity\n")
    
    if extractor.extract_page_8():
        if extractor.create_excel_file():
            print(f"\n✨ SUCCESS! File: {output_file}")
    else:
        print("\n❌ Extraction failed")


if __name__ == "__main__":
    main()
