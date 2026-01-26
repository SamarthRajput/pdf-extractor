"""
PDF Financial Statement Extractor with Standardized Column Schema
Outputs all tables into a SINGLE SHEET with specified column structure
Two options: 1) Page 8 only, 2) Whole PDF
"""

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class FinancialStatementExtractorSingleSheet:
    
    def __init__(self, pdf_path: str, output_path: str = "financial_statements.xlsx"):
        self.pdf_path = pdf_path
        self.output_path = output_path
        self.extracted_tables = []
        self.table_metadata = []
        
        # Standardized columns
        self.standard_columns = [
            'primary_key', 'date_last_updated', 'published_date', 'reported_date',
            'doc_page_num', 'file_page_num', 'table_id', 'country', 'geo_1_id',
            'geo_1_name', 'geo_1_type', 'geo_2_id', 'geo_2_name', 'geo_2_type',
            'dim_4_id', 'dim_4_name', 'dim_3_id', 'dim_3_name', 'dim_2_id',
            'dim_2_name', 'dim_1_id', 'dim_1_name', 'metric_id', 'metric_name',
            'source_metric_id', 'source_metric_name', 'indentation', 'process_flag',
            'base_factor', 'display_power_factor', 'data_frequency', 'aggregation_method',
            'unit', 'unit_type', 'note_id', 'note_reference', 'cumulative_periods',
            'comments', 'check_sum', 'concat', 'formula', '2022', '2023', '2024',
            '2022_check', '2023_check', '2024_check'
        ]
    
    # ============================================================================
    # OPTION 1: EXTRACT ONLY PAGE 8 (Consolidated Statement of Changes in Equity)
    # ============================================================================
    
    def extract_page_8_only(self):
        """Extract only page 8 (Consolidated Statement of Changes in Equity)"""
        print(f"🔄 Opening PDF: {self.pdf_path}")
        print(f"📄 Extracting Page 8 only (Consolidated Statement of Changes in Equity)")
        
        with pdfplumber.open(self.pdf_path) as pdf:
            if len(pdf.pages) < 8:
                print(f"❌ PDF has only {len(pdf.pages)} pages. Page 8 not available!")
                return self.extracted_tables
            
            page_num = 8  # Page 8 in 1-based indexing
            page = pdf.pages[page_num - 1]  # Convert to 0-based indexing
            
            tables = page.extract_tables()
            
            if tables:
                print(f"📊 Found {len(tables)} table(s) on page {page_num}")
                
                for table_idx, table in enumerate(tables):
                    df = self._convert_table_to_dataframe(table)
                    
                    if df is not None and not df.empty:
                        self.extracted_tables.append(df)
                        self.table_metadata.append({
                            'page': page_num,
                            'table_index': table_idx,
                            'rows': len(df),
                            'columns': len(df.columns),
                            'title': f"Consolidated Statement of Changes in Equity - Table {table_idx + 1}"
                        })
                        print(f"✅ Table {len(self.extracted_tables)}: {len(df)} rows × {len(df.columns)} columns")
            else:
                print(f"⚠️ No tables found on page {page_num}")
        
        return self.extracted_tables
    
    # ============================================================================
    # OPTION 2: EXTRACT WHOLE PDF
    # ============================================================================
    
    def extract_all_tables(self):
        """Extract all tables from entire PDF"""
        print(f"🔄 Opening PDF: {self.pdf_path}")
        
        with pdfplumber.open(self.pdf_path) as pdf:
            print(f"📄 Total pages: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()
                
                if tables:
                    print(f"📊 Found {len(tables)} table(s) on page {page_num}")
                    
                    for table_idx, table in enumerate(tables):
                        df = self._convert_table_to_dataframe(table)
                        
                        if df is not None and not df.empty:
                            self.extracted_tables.append(df)
                            self.table_metadata.append({
                                'page': page_num,
                                'table_index': table_idx,
                                'rows': len(df),
                                'columns': len(df.columns)
                            })
                            print(f"✅ Table {len(self.extracted_tables)}: Page {page_num}, {len(df)} rows × {len(df.columns)} columns")
        
        return self.extracted_tables
    
    # ============================================================================
    # HELPER METHODS
    # ============================================================================
    
    def _convert_table_to_dataframe(self, table):
        """Convert raw table to DataFrame"""
        if not table or len(table) < 2:
            return None
        
        try:
            headers = table[0]
            data_rows = table[1:]
            data_rows = [row for row in data_rows if any(cell for cell in row)]
            
            if not data_rows:
                return None
            
            max_cols = len(headers)
            padded_rows = []
            for row in data_rows:
                while len(row) < max_cols:
                    row.append("")
                padded_rows.append(row[:max_cols])
            
            df = pd.DataFrame(padded_rows, columns=headers)
            df = self._clean_dataframe(df)
            
            return df
        except Exception as e:
            print(f"⚠️ Error converting table: {e}")
            return None
    
    def _clean_dataframe(self, df):
        """Clean DataFrame"""
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
        df = df.replace('', None)
        
        return df.reset_index(drop=True)
    
    def convert_numeric_columns(self):
        """Convert numeric columns"""
        for df in self.extracted_tables:
            for col in df.columns:
                df[col] = df[col].apply(lambda x: self._convert_to_numeric(x))
    
    def _convert_to_numeric(self, value):
        """Convert value to numeric if possible"""
        if value is None:
            return value
        
        if isinstance(value, pd.Series):
            return value
        
        value_str = str(value).strip()
        
        if value_str == "" or value_str == "nan":
            return None
        
        # Handle negative in parentheses: (123) -> -123
        if value_str.startswith('(') and value_str.endswith(')'):
            try:
                num = float(value_str[1:-1].replace(',', '').replace(' ', ''))
                return -num
            except:
                return value_str
        
        # Try to convert regular number
        try:
            clean_value = value_str.replace(',', '').replace(' ', '')
            float(clean_value)
            return clean_value
        except:
            return value_str
    
    # ============================================================================
    # CONVERT TO STANDARDIZED SCHEMA
    # ============================================================================
    
    def transform_to_standard_schema(self):
        """
        Transform extracted tables to standardized column schema
        Maps extracted columns to standard schema
        """
        print(f"\n📝 Transforming {len(self.extracted_tables)} table(s) to standardized schema...")
        
        transformed_data = []
        primary_key_counter = 1
        
        for table_idx, (df, metadata) in enumerate(zip(self.extracted_tables, self.table_metadata)):
            print(f"\n📊 Processing Table {table_idx + 1} (Page {metadata['page']})")
            
            # Initialize row data with standard columns
            for row_idx, row in enumerate(df.itertuples(index=False)):
                row_data = {}
                
                # Set standard fields
                row_data['primary_key'] = primary_key_counter
                row_data['date_last_updated'] = datetime.now().strftime('%Y-%m-%d')
                row_data['published_date'] = datetime.now().strftime('%Y-%m-%d')
                row_data['reported_date'] = '2024-12-31'  # Adjust based on your PDF
                row_data['doc_page_num'] = metadata['page']
                row_data['file_page_num'] = metadata['page']
                row_data['table_id'] = f"TABLE_{table_idx + 1:03d}"
                row_data['country'] = 'PL'  # Poland - adjust based on your data
                
                # Geography (optional - fill if available)
                row_data['geo_1_id'] = None
                row_data['geo_1_name'] = None
                row_data['geo_1_type'] = None
                row_data['geo_2_id'] = None
                row_data['geo_2_name'] = None
                row_data['geo_2_type'] = None
                
                # Dimensions - map extracted columns to dimensions
                # Adjust these mappings based on your actual table structure
                extracted_cols = list(row)
                
                # Try to map first few columns to dimensions
                if len(extracted_cols) > 0:
                    row_data['dim_1_id'] = f"DIM1_{table_idx}_{row_idx}"
                    row_data['dim_1_name'] = str(extracted_cols[0])
                else:
                    row_data['dim_1_id'] = None
                    row_data['dim_1_name'] = None
                
                if len(extracted_cols) > 1:
                    row_data['dim_2_id'] = f"DIM2_{table_idx}_{row_idx}"
                    row_data['dim_2_name'] = str(extracted_cols[1])
                else:
                    row_data['dim_2_id'] = None
                    row_data['dim_2_name'] = None
                
                if len(extracted_cols) > 2:
                    row_data['dim_3_id'] = f"DIM3_{table_idx}_{row_idx}"
                    row_data['dim_3_name'] = str(extracted_cols[2])
                else:
                    row_data['dim_3_id'] = None
                    row_data['dim_3_name'] = None
                
                if len(extracted_cols) > 3:
                    row_data['dim_4_id'] = f"DIM4_{table_idx}_{row_idx}"
                    row_data['dim_4_name'] = str(extracted_cols[3])
                else:
                    row_data['dim_4_id'] = None
                    row_data['dim_4_name'] = None
                
                # Metric fields
                row_data['metric_id'] = f"METRIC_{table_idx}_{row_idx}"
                row_data['metric_name'] = f"Value_{row_idx}"
                row_data['source_metric_id'] = f"SOURCE_{table_idx}_{row_idx}"
                row_data['source_metric_name'] = "PDF_Extracted"
                
                # Processing fields
                row_data['indentation'] = 0
                row_data['process_flag'] = 1
                row_data['base_factor'] = 1
                row_data['display_power_factor'] = 0
                row_data['data_frequency'] = 'Annual'
                row_data['aggregation_method'] = 'Sum'
                row_data['unit'] = 'PLN'
                row_data['unit_type'] = 'Thousands'
                
                # Notes
                row_data['note_id'] = None
                row_data['note_reference'] = None
                row_data['cumulative_periods'] = 1
                row_data['comments'] = f"Extracted from page {metadata['page']}"
                
                # Year data - try to extract 2022, 2023, 2024 values
                year_columns = ['2022', '2023', '2024']
                year_check_columns = ['2022_check', '2023_check', '2024_check']
                
                for year_col in year_columns:
                    row_data[year_col] = None
                    row_data[f'{year_col}_check'] = None
                
                # Try to find year columns in extracted data
                if len(extracted_cols) >= 4:
                    row_data['2024'] = extracted_cols[-1] if len(extracted_cols) > 3 else None
                if len(extracted_cols) >= 5:
                    row_data['2023'] = extracted_cols[-2] if len(extracted_cols) > 4 else None
                if len(extracted_cols) >= 6:
                    row_data['2022'] = extracted_cols[-3] if len(extracted_cols) > 5 else None
                
                # Additional fields
                row_data['check_sum'] = None
                row_data['concat'] = None
                row_data['formula'] = None
                
                transformed_data.append(row_data)
                primary_key_counter += 1
        
        # Create DataFrame with standard columns
        result_df = pd.DataFrame(transformed_data, columns=self.standard_columns)
        
        print(f"✅ Transformed {len(result_df)} rows to standardized schema")
        return result_df
    
    def create_single_sheet_excel(self, transformed_df):
        """Create Excel file with all data in a SINGLE SHEET"""
        print(f"\n📝 Creating Excel file with single sheet: {self.output_path}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated_Data"
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Write headers
        for col_idx, header in enumerate(transformed_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Write data
        for row_idx, row in enumerate(transformed_df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                
                # Align based on column type
                col_name = transformed_df.columns[col_idx - 1]
                if col_name in ['2022', '2023', '2024', '2022_check', '2023_check', '2024_check', 
                               'base_factor', 'display_power_factor']:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = left_align
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(transformed_df.columns, 1):
            max_length = max(
                len(str(header)) for header in transformed_df.columns
            )
            if len(transformed_df) > 0:
                col_max = transformed_df.iloc[:, col_idx - 1].astype(str).map(len).max()
                max_length = max(max_length, col_max)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        wb.save(self.output_path)
        print(f"✅ Excel file saved: {self.output_path}")
        print(f"📊 Total rows in sheet: {len(transformed_df)}")
        print(f"📊 Total columns: {len(transformed_df.columns)}")
        
        return self.output_path


# ============================================================================
# MAIN EXECUTION - CHOOSE YOUR OPTION
# ============================================================================

def main_option_1_page_8_only():
    """OPTION 1: Extract ONLY Page 8 (Consolidated Statement of Changes in Equity)"""
    print("\n" + "="*80)
    print("OPTION 1: EXTRACT PAGE 8 ONLY")
    print("="*80)
    
    pdf_file = "2024_Budimex.pdf"
    output_file = "Budimex_Page8_StandardSchema.xlsx"
    
    if not os.path.exists(pdf_file):
        print(f"❌ PDF file not found: {pdf_file}")
        return
    
    extractor = FinancialStatementExtractorSingleSheet(pdf_file, output_file)
    
    # Extract only page 8
    extractor.extract_page_8_only()
    
    # Convert numeric values
    print("\n⚙️ Converting numeric values...")
    extractor.convert_numeric_columns()
    
    # Transform to standard schema
    transformed_df = extractor.transform_to_standard_schema()
    
    # Create single sheet Excel
    extractor.create_single_sheet_excel(transformed_df)
    
    print(f"\n✨ Complete! Output: {output_file}")


def main_option_2_whole_pdf():
    """OPTION 2: Extract WHOLE PDF"""
    print("\n" + "="*80)
    print("OPTION 2: EXTRACT WHOLE PDF")
    print("="*80)
    
    pdf_file = "2024_Budimex.pdf"
    output_file = "Budimex_WholePDF_StandardSchema.xlsx"
    
    if not os.path.exists(pdf_file):
        print(f"❌ PDF file not found: {pdf_file}")
        return
    
    extractor = FinancialStatementExtractorSingleSheet(pdf_file, output_file)
    
    # Extract all tables from entire PDF
    print("\n🚀 Starting extraction process...\n")
    extractor.extract_all_tables()
    
    # Convert numeric values
    print("\n⚙️ Converting numeric values...")
    extractor.convert_numeric_columns()
    
    # Transform to standard schema
    transformed_df = extractor.transform_to_standard_schema()
    
    # Create single sheet Excel
    extractor.create_single_sheet_excel(transformed_df)
    
    print(f"\n✨ Complete! Output: {output_file}")
    print(f"\n📊 Summary:")
    print(f"   - Total tables extracted: {len(extractor.extracted_tables)}")
    print(f"   - Total rows in output: {len(transformed_df)}")
    print(f"   - Standard columns: {len(transformed_df.columns)}")


if __name__ == "__main__":
    # CHOOSE ONE:
    
    # # Option 1: Page 8 only
    # main_option_1_page_8_only()
    
    # Option 2: Whole PDF (uncomment to use)
    main_option_2_whole_pdf()
